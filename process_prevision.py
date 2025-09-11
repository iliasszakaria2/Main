import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def update_pdp(pdp_path: str, prevision_path: str) -> None:
    """Update PDP workbook using data from Prévision EMEA workbook.

    Parameters
    ----------
    pdp_path:
        Path to the PDP Excel workbook.
    prevision_path:
        Path to the Prévision EMEA Excel workbook.
    """
    # --- Load Prévision EMEA and apply filters ---
    df = pd.read_excel(prevision_path, sheet_name="Feuil 1")

    # Normalise fiability rate into [0, 1] range
    fiab = pd.to_numeric(df["Fiability rate"].astype(str).str.replace("%", ""), errors="coerce")
    if fiab.max() > 1:
        fiab = fiab / 100.0

    df = df[(fiab >= 0.60) & (fiab <= 0.90) & (df["EnerOne B-Cab"].notna())]

    df["Livraison"] = pd.to_datetime(df["Livraison"], errors="coerce")
    df["EnerOne B-Cab"] = pd.to_numeric(df["EnerOne B-Cab"], errors="coerce")

    sums = df.groupby("Livraison")["EnerOne B-Cab"].sum()

    # --- Load PDP workbook ---
    wb = load_workbook(pdp_path)
    ws = wb["B-CAB"]

    # Locate the base row and the "Customer Opportunities" row
    base_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "B-CAB L E 0.5C":
            base_row = r
            break
    if base_row is None:
        raise ValueError("'B-CAB L E 0.5C' not found in column A.")

    customer_row = None
    for r in range(base_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Customer Opportunities":
            customer_row = r
            break
    if customer_row is None:
        raise ValueError("'Customer Opportunities' not found below 'B-CAB L E 0.5C'.")

    # Build mapping from date in row 78 to column index
    date_row = 78
    date_to_col = {}
    for c in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=date_row, column=c).value
        if cell_value is None:
            continue
        try:
            date = pd.to_datetime(cell_value).normalize()
        except Exception:
            continue
        date_to_col[date] = c

    # Write sums into the "Customer Opportunities" row
    for date, value in sums.items():
        col = date_to_col.get(pd.Timestamp(date).normalize())
        if col is not None:
            ws.cell(row=customer_row, column=col, value=value)

    wb.save(pdp_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        prog = Path(sys.argv[0]).name
        print(f"Usage: {prog} <PDP.xlsx> <Prevision_EMEA.xlsx>")
        sys.exit(1)
    update_pdp(sys.argv[1], sys.argv[2])
